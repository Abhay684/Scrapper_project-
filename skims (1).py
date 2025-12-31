import asyncio
import random
import re
import urllib.parse
from playwright.async_api import async_playwright
from openpyxl import Workbook
from tqdm.asyncio import tqdm_asyncio

excel_lock = asyncio.Lock()

# ---------------- CONFIG ---------------- #

MAX_CONTEXTS = 3
MAX_CONCURRENT_PDP = 8

COLLECTION_URLS = [
    "https://skims.com/en-in/collections/shapewear",
]

# ---------------- UTILITIES ---------------- #

async def human_pause(a=0.6, b=1.6):
    await asyncio.sleep(random.uniform(a, b))

async def safe_goto(page, url, retries=3):
    for i in range(retries):
        try:
            await page.goto(url, timeout=60000, wait_until="domcontentloaded")
            return True
        except:
            await asyncio.sleep(3 * (i + 1))
    return False

# ---------------- COLLECTION ---------------- #

async def collect_product_links(page):
    links = set()
    stagnant = 0
    prev = 0

    for _ in range(50):
        cards = await page.query_selector_all("div.product-card")
        if cards:
            await cards[-1].scroll_into_view_if_needed()

        for _ in range(random.randint(4, 8)):
            await page.mouse.wheel(0, random.randint(80, 160))
            await page.evaluate(f"window.scrollBy(0,{random.randint(40,90)})")
            await asyncio.sleep(random.uniform(0.15, 0.3))

        anchors = await page.query_selector_all('div.product-card a[href*="/products/"]')
        for a in anchors:
            href = await a.get_attribute("href")
            if href:
                links.add(urllib.parse.urljoin("https://skims.com", href))

        if len(links) == prev:
            stagnant += 1
        else:
            stagnant = 0

        if stagnant >= 4:
            break

        prev = len(links)
        await human_pause(0.8, 1.4)

    return list(links)

# ---------------- REVIEW LOGIC ---------------- #

def parse_months(text):
    if not text:
        return None
    m = re.search(r"(\d+)\s+(day|week|month|year)s?\s+ago", text.lower())
    if not m:
        return None
    n, u = int(m.group(1)), m.group(2)
    return n/30 if u=="day" else n/4 if u=="week" else n if u=="month" else n*12

def safe_int(value):
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    try:
        digits = re.findall(r"\d+", str(value))
        return int(digits[0]) if digits else 0
    except:
        return 0

async def expand_reviews(page):
    for _ in range(60):
        btn = await page.query_selector("span.oke-showMore-button-text")
        if not btn:
            break
        try:
            await btn.click()
            await human_pause(0.8, 1.4)
        except:
            break

async def extract_reviews(page):
    results = []
    try:
        spans = await page.query_selector_all('span:has-text("ago")')
    except:
        return []

    for s in spans:
        try:
            data = await s.evaluate("""
                el => {
                    const card = el.closest('article,li,div');
                    if(!card) return null;
                    const t = sel => card.querySelector(sel)?.innerText || "";
                    return {
                        age: el.innerText,
                        text: t('p'),
                        name: t('.oke-reviewer-name'),
                        pos: t('.oke-helpful-vote-button--positive .oke-helpful-vote-counter'),
                        neg: t('.oke-helpful-vote-button--negative .oke-helpful-vote-counter')
                    }
                }
            """)
        except:
            continue

        m = parse_months(data["age"])
        results.append({
            "within": m is not None and m < 12,
            "age": data["age"],
            "name": data["name"],
            "text": data["text"],
            "pos": safe_int(data.get("pos")),
            "neg": safe_int(data.get("neg"))
        })
    return results

# ---------------- PDP SCRAPER ---------------- #

async def scrape_pdp(semaphore, context, url, ws_products, ws_reviews, workbook):
    async with semaphore:
        page = await context.new_page()
        try:
            if not await safe_goto(page, url):
                return

            try:
                await page.wait_for_selector("h1", timeout=30000)
                await page.evaluate("window.scrollBy(0,1)")
                await page.evaluate("window.scrollBy(0,-1)")
            except:
                return

            await human_pause(1.5, 2.5)

            title = await page.inner_text("h1")
            price = await page.inner_text('span[data-testid="product-current-price"]') if await page.query_selector('span[data-testid="product-current-price"]') else ""
            rating = await page.inner_text("div.oke-sr-rating") if await page.query_selector("div.oke-sr-rating") else "0"
            total_reviews = await page.inner_text("span.oke-sr-count-number") if await page.query_selector("span.oke-sr-count-number") else "0"

            await expand_reviews(page)
            reviews = await extract_reviews(page)

            recent = [r for r in reviews if r["within"]]

            async with excel_lock:
                ws_products.append([
                    title, price, rating, total_reviews,
                    len(recent), sum(r["pos"] for r in recent), url
                ])

                for r in reviews:
                    ws_reviews.append([
                        title, url, r["name"], r["age"],
                        int(r["within"]), r["pos"], r["neg"], r["text"]
                    ])

                workbook.save("skims_products_autosave.xlsx")

            print(f"✅ {title}")

        except Exception as e:
            print(f"⚠️ PDP failed safely: {url} → {e}")

        finally:
            try:
                await page.close()
            except:
                pass

# ---------------- SAFE TASK WRAPPER ---------------- #

async def safe_task(coro):
    try:
        return await coro
    except Exception as e:
        print(f"⚠️ Task failed safely: {e}")
        return None

# ---------------- MAIN ---------------- #

async def main():
    wb = Workbook()

    ws_r = wb.create_sheet("Reviews")
    ws_r.append(["Product","URL","Name","Age","<12M","Yes","No","Text"])

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--disable-dev-shm-usage", "--no-sandbox"]
        )
        contexts = [await browser.new_context() for _ in range(MAX_CONTEXTS)]
        semaphore = asyncio.Semaphore(MAX_CONCURRENT_PDP)

        for i, col in enumerate(COLLECTION_URLS):
            sheet_name = col.split("/collections/")[1].replace("-", " ").title()[:31]

            ws_p = wb.create_sheet(sheet_name)
            ws_p.append(["Name","Price","Rating","Reviews","Recent <12M","Likes <12M","URL"])

            ctx = contexts[i % MAX_CONTEXTS]
            page = await ctx.new_page()
            await safe_goto(page, col)
            await page.wait_for_selector("div.product-card", timeout=30000)
            links = await collect_product_links(page)
            await page.close()

            tasks = [
                safe_task(scrape_pdp(semaphore, ctx, link, ws_p, ws_r, wb))
                for link in links
            ]
            results = await tqdm_asyncio.gather(*tasks)

        await browser.close()

    wb.save("skims_products.xlsx")
    print("🎉 DONE")

if __name__ == "__main__":
    asyncio.run(main())