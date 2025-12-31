"""
BOODY.COM.AU — FULLY UNDETECTED SCRAPER (2025)
Bypasses bot detection → Gets ALL pages
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import *
from selenium.webdriver.common.action_chains import ActionChains
import time
import random
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class UndetectedBoodyScraper:
    def __init__(self):
        self.base_url = "https://www.boody.com.au"
        self.products = []
        self.seen_urls = set()
        self.driver = None
        self.setup_undetected_driver()

    def setup_undetected_driver(self):
        options = Options()
        # REMOVE headless for better detection bypass
        # options.add_argument('--headless=new')  # ← REMOVE THIS LINE
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--disable-extensions')
        options.add_argument('--start-maximized')
        options.add_argument('--disable-infobars')
        options.add_argument('--disable-notifications')
        options.add_argument('--lang=en-US')

        # Critical: Spoof everything
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument('--disable-web-security')
        options.add_argument('--allow-running-insecure-content')

        # Random real user agents
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:130.0) Gecko/20100101 Firefox/130.0"
        ]
        options.add_argument(f'--user-agent={random.choice(user_agents)}')

        self.driver = webdriver.Chrome(options=options)
        
        # FULLY HIDE Selenium
        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
            Object.defineProperty(navigator, 'webdriver', {get: () => false});
            window.navigator.chrome = { runtime: {},  };
            Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
            Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
            """
        })

    def safe_get(self, url, retries=3):
        attempt = 0
        while attempt < retries:
            try:
                self.driver.get(url)
                return True
            except Exception:
                try:
                    self.driver.quit()
                except Exception:
                    pass
                time.sleep(random.uniform(1.5, 3.5))
                try:
                    self.setup_undetected_driver()
                except Exception:
                    pass
                attempt += 1
                time.sleep(random.uniform(1.0, 2.0))
        return False

    def wait_and_click(self, selector, timeout=10):
        for _ in range(timeout):
            try:
                btn = self.driver.find_element(By.CSS_SELECTOR, selector)
                if btn.is_displayed() and btn.is_enabled():
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                    time.sleep(random.uniform(1, 2))
                    btn.click()
                    return True
            except:
                time.sleep(1)
        return False

    def get_all_product_urls(self, collection_url):
        print(f"\nLoading collection: {collection_url}")
        self.safe_get(collection_url)
        time.sleep(random.uniform(6, 8))
        try:
            self._handle_popups_and_region()
        except:
            pass

        all_urls = set()
        
        # Extract products from page 1
        print("Extracting products from page 1...")
        try:
            self._scroll_to_load_products()
        except:
            pass
        page1_urls = self._extract_products_from_current_page()
        all_urls.update(page1_urls)
        print(f"Page 1: {len(page1_urls)} products")
        try:
            self._exhaust_load_more_on_collection()
            more_urls = self._extract_products_from_current_page()
            new_after_load = more_urls - all_urls
            all_urls.update(more_urls)
            if new_after_load:
                print(f"After load-more: +{len(new_after_load)} products")
        except:
            pass
        
        nums = self._get_pagination_numbers()
        pagination_pages = self._find_all_pagination_pages(collection_url) if not nums else {}
        
        # If pagination not detected, try click-through 'Next' buttons (JS-based pagination)
        if not pagination_pages and not nums:
            next_urls = self._scrape_via_next_button(collection_url)
            if next_urls:
                all_urls.update(next_urls)
                print(f"\nTOTAL UNIQUE: {len(all_urls)} products")
                return list(all_urls)
        
        if not pagination_pages:
            print("No pagination found - trying numbered pages")

            # First: try clicking visible numeric buttons (e.g., <li class="mx-1">2</li>) before attempting ?page= URLs
            clicked_urls = self._try_click_numbered_pages()
            if clicked_urls:
                new_after_click = set(clicked_urls) - all_urls
                all_urls.update(clicked_urls)
                print(f"\nTOTAL UNIQUE: {len(all_urls)} products (after clicking numbered buttons; +{len(new_after_click)} new)")
                return list(all_urls)

            # Fallback: attempt loading numbered ?page= URLs (some sites respond to that)
            misses = 0
            for page in range(2, 51):
                try:
                    if '?' in collection_url:
                        test_url = f"{collection_url}&page={page}"
                    else:
                        test_url = f"{collection_url}?page={page}"
                    print(f"Loading page {page}...")
                    self.safe_get(test_url)
                    time.sleep(random.uniform(4, 6))
                    try:
                        self._handle_popups_and_region()
                    except:
                        pass
                    try:
                        self._scroll_to_load_products()
                    except:
                        pass
                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(random.uniform(2, 3))
                    page_urls = self._extract_products_from_current_page()
                    all_urls.update(page_urls)
                    print(f"  Page {page}: {len(page_urls)} products")
                    if len(page_urls) == 0:
                        misses += 1
                        if misses >= 2:
                            break
                    else:
                        misses = 0
                except Exception as e:
                    print(f"  Error on page {page}: {e}")
                    misses += 1
                    if misses >= 2:
                        break
            print(f"\nTOTAL UNIQUE: {len(all_urls)} products")
            return list(all_urls)
        
        if nums:
            print(f"Found {len(nums)} additional pages to scrape")
            for page_num in nums:
                try:
                    print(f"Loading page {page_num}...")
                    success = self._click_pagination_page(page_num)
                    if not success:
                        print(f"  Warning: clicking page {page_num} did not produce a detectable change; continuing to extract current products")
                    time.sleep(random.uniform(3, 5))
                    try:
                        self._scroll_to_load_products()
                    except:
                        pass
                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(random.uniform(2, 3))
                    page_urls = self._extract_products_from_current_page()
                    new_urls = page_urls - all_urls
                    all_urls.update(page_urls)
                    print(f"  Page {page_num}: {len(page_urls)} products ({len(new_urls)} new)")
                except Exception as e:
                    print(f"  Error on page {page_num}: {e}")
                    continue
        else:
            print(f"Found {len(pagination_pages)} additional pages to scrape")
            for page_num, page_url in pagination_pages.items():
                try:
                    print(f"Loading page {page_num}...")
                    self.safe_get(page_url)
                    time.sleep(random.uniform(4, 6))
                    try:
                        self._handle_popups_and_region()
                    except:
                        pass
                    try:
                        self._scroll_to_load_products()
                    except:
                        pass
                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(random.uniform(2, 3))
                    page_urls = self._extract_products_from_current_page()
                    if len(page_urls) == 0:
                        try:
                            alt_url = page_url if '#product-grid' in page_url else (page_url + '#product-grid')
                            self.safe_get(alt_url)
                            time.sleep(random.uniform(3.0, 4.0))
                            try:
                                self._scroll_to_load_products()
                            except:
                                pass
                            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                            time.sleep(random.uniform(2, 3))
                            page_urls = self._extract_products_from_current_page()
                        except:
                            pass
                    new_urls = page_urls - all_urls
                    all_urls.update(page_urls)
                    print(f"  Page {page_num}: {len(page_urls)} products ({len(new_urls)} new)")
                except Exception as e:
                    print(f"  Error on page {page_num}: {e}")
                    continue
            
        
        print(f"\nTOTAL UNIQUE: {len(all_urls)} products from {len(pagination_pages) + 1} pages")
        return list(all_urls)

    def _find_pagination_container(self):
        container = None
        selectors = [
            '[class*="pagination"]',
            '[class*="pager"]',
            'nav[aria-label*="pagination" i]',
            '[role="navigation"]',
            'ul[class*="pagination"]',
            'ul[class*="pages"]',
            '[class*="page-numbers"]',
            '[class*="pager__items"]',
            '[class*="mx-1"]',
        ]
        for sel in selectors:
            try:
                found = self.driver.find_elements(By.CSS_SELECTOR, sel)
                if found:
                    container = found[0]
                    break
            except:
                continue

        # Fallback: find a <ul> whose children include at least two numeric <li> elements
        if not container:
            try:
                uls = self.driver.find_elements(By.TAG_NAME, 'ul')
                for ul in uls:
                    try:
                        li_nums = 0
                        li_children = ul.find_elements(By.TAG_NAME, 'li')
                        for li in li_children:
                            try:
                                if (li.text or '').strip().isdigit():
                                    li_nums += 1
                            except:
                                continue
                        if li_nums >= 2:
                            container = ul
                            break
                    except:
                        continue
            except:
                pass

        return container

    def _get_pagination_numbers(self):
        nums = []
        try:
            cont = self._find_pagination_container()
            elems = []
            if cont:
                try:
                    elems = cont.find_elements(By.CSS_SELECTOR, 'li, a, button')
                except:
                    elems = []
            if not elems:
                try:
                    elems = self.driver.find_elements(By.CSS_SELECTOR, 'li, a, button')
                except:
                    elems = []
            for e in elems:
                try:
                    t = (e.text or '').strip()
                    if t.isdigit():
                        n = int(t)
                        if 2 <= n <= 100:
                            nums.append(n)
                except:
                    continue
            nums = sorted(list(set(nums)))
        except:
            pass
        return nums

    def _click_pagination_page(self, page_num):
        """Attempt to click the page number and wait until products/DOM reflect the new page.
        Returns True if a detectable change occurred, False otherwise."""
        try:
            # snapshot current product links to detect changes
            prior_urls = set()
            try:
                prior_urls = self._extract_products_from_current_page()
            except:
                prior_urls = set()

            before = None
            try:
                containers = []
                for sel in ['#product-grid', '.product-grid', '.collection__products', '.grid--uniform']:
                    try:
                        found = self.driver.find_elements(By.CSS_SELECTOR, sel)
                        if found:
                            containers = found
                            break
                    except:
                        continue
                if containers:
                    before = containers[0]
            except:
                pass

            target = None
            cont = self._find_pagination_container()
            pool = []
            try:
                if cont:
                    try:
                        pool = cont.find_elements(By.CSS_SELECTOR, 'a, button, li')
                    except:
                        pool = []
                if not pool:
                    try:
                        pool = self.driver.find_elements(By.CSS_SELECTOR, 'a, button, li')
                    except:
                        pool = []
                for e in pool:
                    try:
                        t = (e.text or '').strip()
                        if t == str(page_num) and e.is_displayed():
                            target = e
                            break
                    except:
                        continue
                if not target:
                    try:
                        xpath_elems = self.driver.find_elements(By.XPATH, f"//li[normalize-space()='{page_num}'] | //a[normalize-space()='{page_num}'] | //button[normalize-space()='{page_num}']")
                        for e in xpath_elems:
                            try:
                                if e.is_displayed():
                                    target = e
                                    break
                            except:
                                continue
                    except:
                        pass

                if not target:
                    # nothing to click — try a fallback for sites that use li.mx-1 style pagination buttons
                    try:
                        # Try a common site-specific XPath (as provided)
                        fallback_xpath = "//li[@class='mx-1 flex h-8 w-8 flex-col flex-wrap items-center justify-center bg-grey-100 text-grey-700 cursor-pointer'][1]"
                        elems = self.driver.find_elements(By.XPATH, fallback_xpath)
                        if elems:
                            fallback = elems[0]
                            try:
                                if fallback.is_displayed():
                                    print("  Fallback: clicking special mx-1 li element (XPath match) as page button")
                                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", fallback)
                                    time.sleep(random.uniform(0.6, 1.2))
                                    try:
                                        self.driver.execute_script("arguments[0].click();", fallback)
                                    except:
                                        try:
                                            fallback.click()
                                        except:
                                            pass
                                    # wait & detect change
                                    for _ in range(10):
                                        time.sleep(0.4)
                                        try:
                                            current = self._extract_products_from_current_page()
                                            if current - prior_urls:
                                                print(f"  Fallback click changed content for page {page_num}")
                                                return True
                                        except:
                                            pass
                                        try:
                                            if self._is_page_active(page_num):
                                                print(f"  Fallback click changed active page to {page_num}")
                                                return True
                                        except:
                                            pass
                            except:
                                pass
                    except:
                        pass
                    return False

                # If the target is an <li> (e.g. <li class="mx-1 ...">1</li>), try to find a clickable child inside it
                try:
                    tag = target.tag_name
                except:
                    tag = 'unknown'

                # Try to replace target with a clickable child if available
                try:
                    if tag.lower() == 'li':
                        clickable_child = None
                        try:
                            # prefer anchors and buttons
                            for child in target.find_elements(By.CSS_SELECTOR, 'a, button, [role="button"]'):
                                try:
                                    if child.is_displayed():
                                        clickable_child = child
                                        break
                                except:
                                    continue
                            # fallback: any child with the same visible text
                            if not clickable_child:
                                txt = (target.text or '').strip()
                                for child in target.find_elements(By.CSS_SELECTOR, '*'):
                                    try:
                                        if (child.text or '').strip() == txt and child.is_displayed():
                                            clickable_child = child
                                            break
                                    except:
                                        continue
                        except:
                            clickable_child = None

                        if clickable_child:
                            target = clickable_child
                            tag = target.tag_name
                except:
                    pass

                try:
                    href = target.get_attribute('href') or ''
                except:
                    href = ''
                try:
                    cls = (target.get_attribute('class') or '')
                except:
                    cls = ''
                try:
                    text = (target.text or '').strip()[:120]
                except:
                    text = ''
                print(f"  Clicking target -> tag={tag}, text='{text}', href='{href[:200]}', class='{cls[:120]}'")

                # Try multiple click strategies to trigger JS-driven pagination
                tried = []
                success = False

                # 1) dispatch a DOM MouseEvent click
                try:
                    self.driver.execute_script("arguments[0].dispatchEvent(new MouseEvent('click',{bubbles:true,cancelable:true}));", target)
                    tried.append('dispatchEvent')
                except Exception as e:
                    tried.append(f'dispatchEvent_err:{e}')

                # wait & check
                for _ in range(6):
                    time.sleep(0.4)
                    try:
                        current = self._extract_products_from_current_page()
                        if current - prior_urls:
                            success = True
                            break
                    except:
                        pass
                    try:
                        if self._is_page_active(page_num):
                            success = True
                            break
                    except:
                        pass

                if not success:
                    # 2) try jQuery trigger if available
                    try:
                        self.driver.execute_script("if (window.jQuery) { window.jQuery(arguments[0]).trigger('click'); }", target)
                        tried.append('jquery_trigger')
                    except Exception as e:
                        tried.append(f'jquery_err:{e}')

                    for _ in range(6):
                        time.sleep(0.4)
                        try:
                            current = self._extract_products_from_current_page()
                            if current - prior_urls:
                                success = True
                                break
                        except:
                            pass
                        try:
                            if self._is_page_active(page_num):
                                success = True
                                break
                        except:
                            pass

                if not success:
                    # 3) use ActionChains click
                    try:
                        actions = ActionChains(self.driver)
                        actions.move_to_element(target).pause(random.uniform(0.1,0.4)).click(target).perform()
                        tried.append('actionchains')
                    except Exception as e:
                        tried.append(f'action_err:{e}')

                    for _ in range(6):
                        time.sleep(0.4)
                        try:
                            current = self._extract_products_from_current_page()
                            if current - prior_urls:
                                success = True
                                break
                        except:
                            pass
                        try:
                            if self._is_page_active(page_num):
                                success = True
                                break
                        except:
                            pass

                # 4) Fallback: attempt stepwise 'next' clicks until we detect desired page
                if not success:
                    print(f"  Direct click didn't change content (attempts: {tried}); trying stepwise Next clicks to reach page {page_num}")
                    steps = 0
                    while steps < max(6, page_num):
                        steps += 1
                        # find next button
                        next_btn = None
                        for selector in ['a[rel="next"]', 'a[aria-label*="next" i]', 'button[aria-label*="next" i]', 'a[class*="next" i]']:
                            try:
                                cands = self.driver.find_elements(By.CSS_SELECTOR, selector)
                                for c in cands:
                                    try:
                                        if c.is_displayed() and 'disabled' not in (c.get_attribute('class') or '').lower():
                                            next_btn = c
                                            break
                                    except:
                                        continue
                                if next_btn:
                                    break
                            except:
                                continue
                        if not next_btn:
                            break

                        try:
                            self.driver.execute_script("arguments[0].dispatchEvent(new MouseEvent('click',{bubbles:true,cancelable:true}));", next_btn)
                        except:
                            try:
                                ActionChains(self.driver).move_to_element(next_btn).click(next_btn).perform()
                            except:
                                try:
                                    next_btn.click()
                                except:
                                    break

                        # wait and check
                        for _ in range(6):
                            time.sleep(0.4)
                            try:
                                current = self._extract_products_from_current_page()
                                if current - prior_urls:
                                    if self._is_page_active(page_num) or len(current - prior_urls) > 0:
                                        success = True
                                        break
                            except:
                                pass
                            try:
                                if self._is_page_active(page_num):
                                    success = True
                                    break
                            except:
                                pass
                        if success:
                            print(f"  Reached page {page_num} after {steps} next clicks")
                            break

                if success:
                    print(f"  Click changed content for page {page_num}")
                else:
                    print(f"  Click did not change content for page {page_num} after all attempts (tried: {tried})")

                # final sanity wait for product elements
                try:
                    WebDriverWait(self.driver, 6).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href*="/products/"]'))
                    )
                except:
                    pass

                return bool(success)
            except:
                return False
        except:
            return False
    
    def _extract_products_from_current_page(self):
        """Extract all product URLs from current page"""
        urls = set()
        elements = []
        try:
            containers = []
            container_selectors = [
                '.product-grid',
                '[class*="collection" i] .grid',
                '[data-section-id*="collection" i]',
                '.collection__products',
                '.grid--uniform',
                '.product-list',
                '#product-grid',
                '[id*="product" i]',
                '[class*="product" i]'
            ]
            for sel in container_selectors:
                try:
                    found = self.driver.find_elements(By.CSS_SELECTOR, sel)
                    if found:
                        containers = found
                        break
                except:
                    continue
            if containers:
                for c in containers:
                    try:
                        elements.extend(c.find_elements(By.CSS_SELECTOR, 'a[href^="/products/"], a[href*="/products/"]'))
                    except:
                        continue
            else:
                elements = self.driver.find_elements(By.CSS_SELECTOR, 'a[href^="/products/"], a[href*="/products/"]')

            if not elements:
                try:
                    WebDriverWait(self.driver, 8).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href*="/products/"]'))
                    )
                    elements = self.driver.find_elements(By.CSS_SELECTOR, 'a[href*="/products/"]')
                except:
                    pass

            for link in elements:
                try:
                    if not link.is_displayed():
                        continue
                    href = link.get_attribute('href')
                    if not href:
                        continue
                    clean = href.split('?')[0].split('#')[0].rstrip('/')
                    if '/products/' in clean and clean.startswith('http'):
                        tail = clean.split('/')[-1]
                        if 'gift-card' in tail or 'gift' in tail:
                            continue
                        urls.add(clean)
                except:
                    continue
        except:
            pass
        return urls

    def _is_page_active(self, page_num):
        try:
            cont = self._find_pagination_container()
            pool = []
            if cont:
                try:
                    pool = cont.find_elements(By.CSS_SELECTOR, 'a, button, li')
                except:
                    pool = []
            if not pool:
                try:
                    pool = self.driver.find_elements(By.CSS_SELECTOR, 'a, button, li')
                except:
                    pool = []
            for e in pool:
                try:
                    t = (e.text or '').strip()
                    if t == str(page_num):
                        ac = (e.get_attribute('aria-current') or '').lower()
                        cls = (e.get_attribute('class') or '').lower()
                        if ac in ['page', 'true'] or any(x in cls for x in ['active', 'current', 'is-active']):
                            return True
                except:
                    continue
        except:
            pass
        return False
    
    def _find_all_pagination_pages(self, collection_url):
        """Find all pagination page numbers and URLs"""
        pages = {}
        
        try:
            # Scroll to pagination area
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            
            # Method 1: Find numbered buttons (2, 3, 4, etc.)
            try:
                # Find pagination container
                pagination_container = None
                container_selectors = [
                    '[class*="pagination"]',
                    '[class*="pager"]',
                    'nav[aria-label*="pagination" i]',
                    '[role="navigation"]',
                ]
                
                for selector in container_selectors:
                    try:
                        containers = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        if containers:
                            pagination_container = containers[0]
                            break
                    except:
                        continue
                
                # Find all clickable elements with numbers
                if pagination_container:
                    numbered_elements = pagination_container.find_elements(By.CSS_SELECTOR, 'a, button, [role="button"]')
                else:
                    # Fallback: search entire page
                    numbered_elements = self.driver.find_elements(By.CSS_SELECTOR, 'a, button')
                
                for elem in numbered_elements:
                    try:
                        text = elem.text.strip()
                        # Check if it's a number (page number)
                        if text.isdigit():
                            page_num = int(text)
                            if 2 <= page_num <= 100:  # Skip page 1, reasonable max
                                href = elem.get_attribute('href')
                                
                                # If no href, construct URL
                                if not href or href == '#' or href == 'javascript:void(0)':
                                    if '?' in collection_url:
                                        href = f"{collection_url}&page={page_num}"
                                    else:
                                        href = f"{collection_url}?page={page_num}"
                                
                                # Normalize URL
                                if href.startswith('/'):
                                    href = self.base_url + href
                                
                                if href and href.startswith('http'):
                                    pages[page_num] = href
                                    print(f"  Found page button: {page_num} -> {href}")
                    except:
                        continue
            except Exception as e:
                print(f"  Error finding numbered buttons: {e}")
            
            # Method 1b: Detect non-clickable page numbers in <li> and construct URLs
            try:
                li_numbers = []
                li_elements = []
                if 'pagination_container' in locals() and pagination_container:
                    try:
                        li_elements = pagination_container.find_elements(By.CSS_SELECTOR, 'li')
                    except:
                        li_elements = []
                if not li_elements:
                    try:
                        li_elements = self.driver.find_elements(By.CSS_SELECTOR, 'li')
                    except:
                        li_elements = []
                for li in li_elements:
                    try:
                        txt = (li.text or '').strip()
                        if txt.isdigit():
                            n = int(txt)
                            if 2 <= n <= 100:
                                li_numbers.append(n)
                    except:
                        continue
                if li_numbers:
                    max_page = max(li_numbers)
                    for page in range(2, max_page + 1):
                        if page not in pages:
                            if '?' in collection_url:
                                url = f"{collection_url}&page={page}"
                            else:
                                url = f"{collection_url}?page={page}"
                            pages[page] = url
            except:
                pass

            # Method 2: Find ">" next button and discover more pages
            try:
                # Look for ">" or next button
                next_selectors = [
                    'a[aria-label*="next" i]',
                    'button[aria-label*="next" i]',
                    'a[class*="next" i]',
                    'button[class*="next" i]',
                ]
                
                next_btn = None
                for selector in next_selectors:
                    try:
                        btn = self.driver.find_element(By.CSS_SELECTOR, selector)
                        if btn.is_displayed():
                            btn_class = (btn.get_attribute('class') or '').lower()
                            if 'disabled' not in btn_class:
                                next_btn = btn
                                break
                    except:
                        continue
                
                # Also look for ">" symbol
                if not next_btn:
                    try:
                        all_links = self.driver.find_elements(By.XPATH, "//a[contains(text(), '>')] | //button[contains(text(), '>')]")
                        for link in all_links:
                            text = link.text.strip()
                            if text == '>' or text == '»':
                                btn_class = (link.get_attribute('class') or '').lower()
                                if 'disabled' not in btn_class:
                                    next_btn = link
                                    break
                    except:
                        pass
                
                if next_btn:
                    next_href = next_btn.get_attribute('href')
                    if next_href:
                        if next_href.startswith('/'):
                            next_href = self.base_url + next_href
                        
                        # Extract page number from next button
                        match = re.search(r'[?&/]page[=/](\d+)', next_href.lower())
                        if match:
                            next_page = int(match.group(1))
                            pages[next_page] = next_href
                            print(f"  Found 'Next' button (>) -> page {next_page}")
                            
                            # Try to discover more pages by constructing URLs
                            # If next goes to page 2, try pages 3, 4, 5, etc.
                            for page in range(next_page + 1, next_page + 10):  # Try up to 10 more pages
                                if '?' in collection_url:
                                    test_url = f"{collection_url}&page={page}"
                                else:
                                    test_url = f"{collection_url}?page={page}"
                                if page not in pages:
                                    pages[page] = test_url
            except Exception as e:
                print(f"  Error finding next button: {e}")
            
            # Method 2b: Use rel="next" link in head
            try:
                next_links = self.driver.find_elements(By.CSS_SELECTOR, 'link[rel="next"]')
                for nl in next_links:
                    href = nl.get_attribute('href')
                    if href:
                        if href.startswith('/'):
                            href = self.base_url + href
                        m = re.search(r'[?&/]page[=/](\d+)', href.lower())
                        if m:
                            p = int(m.group(1))
                            pages[p] = href
                            for page in range(p + 1, p + 6):
                                if '?' in href:
                                    base = href.split('?')[0]
                                    test_url = f"{base}?page={page}"
                                else:
                                    test_url = re.sub(r'(page[=/]\d+)', f'page={page}', href)
                                if page not in pages:
                                    pages[page] = test_url
                        else:
                            pages[max(pages.keys()) + 1 if pages else 2] = href
            except:
                pass

            # Method 3: Try to find max page from page source
            if not pages:
                try:
                    page_source = self.driver.page_source
                    patterns = [
                        r'page\s+(\d+)\s+of\s+(\d+)',
                        r'"totalPages"\s*:\s*(\d+)',
                        r'"pageCount"\s*:\s*(\d+)',
                        r'data-total-pages=["\'](\d+)',
                    ]
                    
                    max_page = 1
                    for pattern in patterns:
                        matches = re.findall(pattern, page_source, re.IGNORECASE)
                        for match in matches:
                            if isinstance(match, tuple):
                                for num in match:
                                    try:
                                        page_num = int(num)
                                        if page_num > max_page:
                                            max_page = page_num
                                    except:
                                        pass
                            else:
                                try:
                                    page_num = int(match)
                                    if page_num > max_page:
                                        max_page = page_num
                                except:
                                    pass
                    
                    if max_page > 1:
                        print(f"  Found max page {max_page} in page source")
                        for page in range(2, max_page + 1):
                            if '?' in collection_url:
                                url = f"{collection_url}&page={page}"
                            else:
                                url = f"{collection_url}?page={page}"
                            if page not in pages:
                                pages[page] = url
                except:
                    pass
            
            # Sort pages by number
            pages = dict(sorted(pages.items()))
            
        except Exception as e:
            print(f"  Error finding pagination: {e}")
        
        return pages

    def _try_click_numbered_pages(self, max_pages=50):
        """Try clicking visible numeric page buttons (li/a/button) for pages 2..max_pages.
        Returns a set of discovered product URLs from pages that were successfully navigated to (empty set if none)."""
        discovered = set()
        consecutive_misses = 0
        for page in range(2, max_pages + 1):
            try:
                # find candidate elements matching the number
                candidates = []
                cont = self._find_pagination_container()
                if cont:
                    try:
                        candidates = cont.find_elements(By.XPATH, f".//li[normalize-space()='{page}'] | .//a[normalize-space()='{page}'] | .//button[normalize-space()='{page}']")
                    except:
                        candidates = []
                if not candidates:
                    try:
                        candidates = self.driver.find_elements(By.XPATH, f"//li[normalize-space()='{page}'] | //a[normalize-space()='{page}'] | //button[normalize-space()='{page}']")
                    except:
                        candidates = []

                clicked = False
                for c in candidates:
                    try:
                        if not c.is_displayed():
                            continue
                        print(f"Trying numeric click for page {page} (element tag={c.tag_name}, text='{(c.text or '').strip()}')")
                        # delegate to generic click handler which checks for content change
                        success = self._click_pagination_page(page)
                        clicked = True
                        if success:
                            try:
                                # wait a moment then collect page urls
                                time.sleep(random.uniform(1.2, 2.0))
                                self._scroll_to_load_products()
                            except:
                                pass
                            page_urls = self._extract_products_from_current_page()
                            if page_urls:
                                print(f"  Page {page}: {len(page_urls)} products (via numeric click)")
                                # accumulate discovered URLs
                                discovered.update(page_urls)
                                consecutive_misses = 0
                            else:
                                consecutive_misses += 1
                        else:
                            print(f"  Click for page {page} did not change content")
                            consecutive_misses += 1
                        break
                    except:
                        continue

                if not clicked:
                    consecutive_misses += 1

                if consecutive_misses >= 3:
                    break
            except Exception as e:
                print(f"  Error while trying numeric page {page}: {e}")
                consecutive_misses += 1
                if consecutive_misses >= 3:
                    break
        return discovered

    def _scrape_via_next_button(self, collection_url):
        """Paginate by repeatedly clicking or following the 'next' control and collecting product links."""
        pages = set()
        max_steps = 200
        for step in range(max_steps):
            try:
                # ensure products are loaded on current page
                try:
                    self._scroll_to_load_products()
                except:
                    pass

                page_urls = self._extract_products_from_current_page()
                if page_urls:
                    new = page_urls - pages
                    print(f"  Page {step+1}: {len(page_urls)} products ({len(new)} new)")
                    pages.update(page_urls)
                else:
                    print(f"  Page {step+1}: no product links found")

                # locate a next button/link
                next_btn = None
                selectors = [
                    'a[rel="next"]',
                    'a[aria-label*="next" i]',
                    'button[aria-label*="next" i]',
                    'a[class*="next" i]',
                    'button[class*="next" i]'
                ]
                for sel in selectors:
                    try:
                        candidates = self.driver.find_elements(By.CSS_SELECTOR, sel)
                        for c in candidates:
                            cls = (c.get_attribute('class') or '').lower()
                            if c.is_displayed() and 'disabled' not in cls:
                                next_btn = c
                                break
                        if next_btn:
                            break
                    except:
                        continue

                if not next_btn:
                    # fallback: look for > or » or text 'next' or numeric li siblings like ui style buttons
                    try:
                        candidates = self.driver.find_elements(By.XPATH, "//a[contains(text(),'>')] | //a[contains(text(),'»')] | //a[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'next')] | //button[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'next')] | //li[contains(@class,'mx-1')] | //li[normalize-space()=string(number(normalize-space(.)))]")
                        for c in candidates:
                            try:
                                cls = (c.get_attribute('class') or '').lower()
                                if c.is_displayed() and 'disabled' not in cls:
                                    next_btn = c
                                    break
                            except:
                                continue
                    except:
                        pass

                if not next_btn:
                    print("  No next button found — stopping click-next pagination")
                    break

                # attempt to navigate via href first, otherwise click
                href = None
                try:
                    href = next_btn.get_attribute('href')
                except:
                    href = None
                if href and href.startswith('/'):
                    href = self.base_url + href

                if href and href.startswith('http'):
                    # prefer direct navigation when href is available
                    try:
                        self.safe_get(href)
                    except:
                        try:
                            self.driver.execute_script("arguments[0].click();", next_btn)
                        except:
                            try:
                                next_btn.click()
                            except:
                                break
                else:
                    try:
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_btn)
                        time.sleep(random.uniform(0.6, 1.2))
                        # If the next_btn is an <li />, try to click a child anchor/button instead
                        try:
                            if next_btn.tag_name.lower() == 'li':
                                child = None
                                try:
                                    for ch in next_btn.find_elements(By.CSS_SELECTOR, 'a, button, [role="button"]'):
                                        if ch.is_displayed():
                                            child = ch
                                            break
                                except:
                                    child = None
                                if child:
                                    self.driver.execute_script("arguments[0].click();", child)
                                else:
                                    self.driver.execute_script("arguments[0].click();", next_btn)
                            else:
                                self.driver.execute_script("arguments[0].click();", next_btn)
                        except Exception:
                            try:
                                next_btn.click()
                            except:
                                print("  Could not click next button — stopping")
                                break
                    except Exception:
                        try:
                            next_btn.click()
                        except:
                            print("  Could not click next button — stopping")
                            break

                # Wait for page products to appear
                try:
                    WebDriverWait(self.driver, 8).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href*="/products/"]'))
                    )
                except:
                    time.sleep(random.uniform(1.5,2.5))

            except Exception as e:
                print(f"  Error during click-next loop: {e}")
                break

        return pages

    def get_product_details(self, url):
        try:
            self.safe_get(url)
            time.sleep(random.uniform(4, 6))

            data = {'url': url}

            # Name
            try:
                data['name'] = self.driver.find_element(By.CSS_SELECTOR, 'h1').text.strip()
            except:
                data['name'] = url.split('/')[-1].replace('-', ' ').title()

            try:
                scripts = self.driver.find_elements(By.CSS_SELECTOR, 'script[type="application/ld+json"]')
                product_ld = None
                for s in scripts:
                    try:
                        obj = json.loads(s.get_attribute('innerHTML'))
                    except:
                        continue
                    candidates = []
                    if isinstance(obj, list):
                        candidates = obj
                    elif isinstance(obj, dict) and obj.get('@graph') and isinstance(obj['@graph'], list):
                        candidates = obj['@graph']
                    else:
                        candidates = [obj]
                    for item in candidates:
                        t = item.get('@type')
                        if (t == 'Product') or (isinstance(t, list) and 'Product' in t):
                            product_ld = item
                            break
                    if product_ld:
                        break
                if product_ld:
                    offers = product_ld.get('offers')
                    offer = None
                    if isinstance(offers, list) and offers:
                        offer = offers[0]
                    elif isinstance(offers, dict):
                        offer = offers
                    if offer:
                        pr = offer.get('price') or (offer.get('priceSpecification') or {}).get('price')
                        if pr:
                            try:
                                data['price'] = f"${float(str(pr).replace(',', '')):.2f}"
                            except:
                                data['price'] = str(pr)
                    aggr = product_ld.get('aggregateRating') or {}
                    rv = aggr.get('ratingValue') or aggr.get('rating')
                    rc = aggr.get('ratingCount')
                    rvc = aggr.get('reviewCount')
                    if rv is not None:
                        try:
                            data['rating'] = round(float(str(rv)), 2)
                        except:
                            data['rating'] = str(rv)
                    if rc is not None:
                        try:
                            data['rating_count'] = int(str(rc).replace(',', ''))
                        except:
                            data['rating_count'] = str(rc)
                    if rvc is not None:
                        try:
                            data['review_count'] = int(str(rvc).replace(',', ''))
                        except:
                            data['review_count'] = str(rvc)
            except:
                pass
            if 'price' not in data:
                try:
                    el = self.driver.find_element(By.CSS_SELECTOR, 'meta[itemprop="price"]')
                    val = el.get_attribute('content')
                    if val:
                        data['price'] = f"${float(str(val).replace(',', '')):.2f}"
                except:
                    try:
                        el = self.driver.find_element(By.CSS_SELECTOR, 'meta[property="product:price:amount"]')
                        val = el.get_attribute('content')
                        if val:
                            data['price'] = f"${float(str(val).replace(',', '')):.2f}"
                    except:
                        try:
                            price_el = self.driver.find_element(By.CSS_SELECTOR, '.price, .price__current, .price-item--regular, .money')
                            m = re.search(r'\$[0-9,]+\.?[0-9]*', price_el.text)
                            data['price'] = m.group() if m else "N/A"
                        except:
                            data['price'] = "N/A"
            if 'rating' not in data:
                try:
                    el = self.driver.find_element(By.CSS_SELECTOR, '[itemprop="ratingValue"], meta[itemprop="ratingValue"], [data-rating], .okeReviews-reviewsSummary-rating, .yotpo-stars')
                    txt = el.get_attribute('content') or el.get_attribute('data-rating') or el.get_attribute('title') or el.text
                    m = re.search(r'[0-9]+\.?[0-9]*', txt or '')
                    data['rating'] = round(float(m.group()), 2) if m else None
                except:
                    data['rating'] = None
            if 'review_count' not in data:
                try:
                    el = self.driver.find_element(By.CSS_SELECTOR, '[itemprop="reviewCount"], .reviews-count, [class*="review-count"], .okeReviews-reviewsSummary-count, .yotpo-review-count')
                    txt = el.get_attribute('content') or el.text
                    m = re.search(r'[0-9,]+', txt or '')
                    data['review_count'] = int(m.group().replace(',', '')) if m else None
                except:
                    data['review_count'] = None
            if 'rating_count' not in data:
                data['rating_count'] = data.get('review_count')

            return data
        except:
            return None

    def _exhaust_load_more_on_collection(self):
        tries = 0
        max_clicks = 30
        while tries < max_clicks:
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.uniform(1.5, 2.5))
            btn = None
            selectors = [
                'button[class*="load" i]',
                'a[class*="load" i]',
                'button[class*="more" i]',
                'a[class*="more" i]',
                'button[class*="show" i]',
                'a[class*="show" i]'
            ]
            for sel in selectors:
                try:
                    cand = self.driver.find_elements(By.CSS_SELECTOR, sel)
                    for c in cand:
                        t = (c.text or '').strip().lower()
                        cls = (c.get_attribute('class') or '').lower()
                        if any(x in t for x in ['load more', 'show more', 'view more', 'more products']) and 'disabled' not in cls:
                            btn = c
                            break
                    if btn:
                        break
                except:
                    continue
            if not btn:
                break
            try:
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                time.sleep(random.uniform(0.8, 1.4))
                btn.click()
                time.sleep(random.uniform(2.0, 3.0))
            except:
                break
            tries += 1

    def _scroll_to_load_products(self):
        last_count = 0
        stable = 0
        for _ in range(40):
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.uniform(1.2, 2.0))
            current = len(self._extract_products_from_current_page())
            if current <= last_count:
                stable += 1
                if stable >= 3:
                    break
            else:
                stable = 0
                last_count = current

    def _handle_popups_and_region(self):
        selectors = [
            '#onetrust-accept-btn-handler',
            '[data-accept-cookie]',
            '.cookie__accept',
            '.accept-cookies',
        ]
        for sel in selectors:
            try:
                btns = self.driver.find_elements(By.CSS_SELECTOR, sel)
                for b in btns:
                    if b.is_displayed():
                        try:
                            self.driver.execute_script("arguments[0].click();", b)
                            time.sleep(0.5)
                        except:
                            pass
            except:
                continue
        try:
            modals = self.driver.find_elements(By.CSS_SELECTOR, '[role="dialog"], .modal, .popup')
            for m in modals:
                try:
                    close_btns = m.find_elements(By.CSS_SELECTOR, '[aria-label*="close" i], .close, .modal__close, button[type="button"]')
                    for cb in close_btns:
                        if cb.is_displayed():
                            self.driver.execute_script("arguments[0].click();", cb)
                            time.sleep(0.5)
                            break
                except:
                    continue
        except:
            pass

    def scrape_collection(self, url, name):
        urls = self.get_all_product_urls(url)
        products = []
        for i, u in enumerate(urls):
            if u in self.seen_urls:
                continue
            print(f"[{i+1}/{len(urls)}] {u.split('/')[-1][:40]}...")
            prod = self.get_product_details(u)
            if prod:
                prod['collection'] = name
                products.append(prod)
                self.seen_urls.add(u)
            time.sleep(random.uniform(0.8, 2))
        return products

    def run(self):
        collections = {
            "Women's Bras": "https://www.boody.com.au/collections/womens-bras",
            "Women's Underwear": "https://www.boody.com.au/collections/womens-underwear",
            "Lounge & Sleep": "https://www.boody.com.au/collections/womens-lounge-sleep",
            "Women's Clothing": "https://www.boody.com.au/collections/womens-clothing",
        }

        all_products = []
        for name, url in collections.items():
            print(f"\n{'='*90}")
            print(f"SCRAPING: {name}")
            print(f"{'='*90}")
            prods = self.scrape_collection(url, name)
            all_products.extend(prods)
            self.products = all_products

            # Save progress
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Price", "Rating", "Rating Count", "Review Count", "Collection", "URL"])
            for p in all_products:
                ws.append([
                    p.get('name'),
                    p.get('price'),
                    p.get('rating'),
                    p.get('rating_count'),
                    p.get('review_count'),
                    p.get('collection'),
                    p.get('url')
                ])
            wb.save(f"boody_progress_{name.replace(' ', '_')}.xlsx")
            print(f"Saved {len(prods)} from {name}")

        # Final save
        wb = Workbook()
        ws = wb.active
        ws.append(["Product Name", "Price", "Rating", "Rating Count", "Review Count", "Collection", "URL"])
        for p in all_products:
            ws.append([
                p.get('name'),
                p.get('price'),
                p.get('rating'),
                p.get('rating_count'),
                p.get('review_count'),
                p.get('collection'),
                p.get('url')
            ])
        wb.save("BOODY_FULL_COMPLETE.xlsx")
        print(f"\nALL DONE! Total: {len(all_products)} products")

        self.driver.quit()

# RUN IT
if __name__ == "__main__":
    scraper = UndetectedBoodyScraper()
    scraper.run()
