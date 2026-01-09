from playwright.sync_api import sync_playwright
import time, random
from openpyxl import Workbook, load_workbook
import re
import csv
import os
from datetime import datetime, timedelta


def scrape_mns_myntra():
    # Add one or more listing URLs here. All products from all URLs will be saved
    # into the same CSV and the same Excel sheet.
    urls = [
        "https://www.myntra.com/lingerie?rawQuery=Lingerie",
        
        
    ]

    scrape_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base_name = "myntra_van_heusen_hipster"
    csv_name = f"{base_name}.csv"
    xlsx_name = f"{base_name}.xlsx"

    csv_headers = [
        "Brand Name",
        "Full Name",
        "Price",
        "Product Rating",
        "Customer Reviews Count",
        "Product URL",
        "Customer Reviews (Last 12 Months) Count",
        "2020",
        "2021",
        "2022",
        "2023",
        "2024",
        "2025",
        "2026",
        "S",
        "M",
        "L",
        "XL",
        "XXL",
        "3XL",
        "4XL",
        "Scrape Timestamp",
    ]

    results = []

    if (not os.path.exists(csv_name)) or (os.path.exists(csv_name) and os.path.getsize(csv_name) == 0):
        with open(csv_name, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(csv_headers)

    def _is_myntra_oops_page(p) -> bool:
        try:
            # Myntra error page typically renders: "Oops! Something went wrong." + a "Refresh" button.
            if p.locator(r"text=/Oops!?\s*Something\s*went\s*wrong\.?/i").count() > 0:
                return True
            if p.locator(r"text=/Something\s+went\s+wrong/i").count() > 0 and p.locator(r"text=/Refresh/i").count() > 0:
                return True
        except Exception:
            return False
        return False

    def _recover_from_myntra_oops(p, target_url: str | None = None, max_attempts: int = 6) -> bool:
        # Best-effort recovery: click Refresh if present, then reload/go-to the target URL.
        # Returns True if page no longer looks like the oops page.
        target_url = (target_url or "").strip() or None
        for _ in range(max_attempts):
            if not _is_myntra_oops_page(p):
                return True
            try:
                btn = p.locator(":is(button,a,div,span):has-text('Refresh'), text=/^Refresh$/i").first
                if btn and btn.count() > 0:
                    try:
                        btn.scroll_into_view_if_needed(timeout=1500)
                    except Exception:
                        pass
                    try:
                        btn.click(timeout=2500)
                    except Exception:
                        pass
            except Exception:
                pass

            try:
                # Manual refresh typically re-navigates to the same URL; do the same when we can.
                if target_url:
                    p.goto(target_url, timeout=90000, wait_until="domcontentloaded")
                else:
                    p.reload(timeout=90000, wait_until="domcontentloaded")
            except Exception:
                pass

            try:
                p.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            time.sleep(0.7 + random.uniform(0.15, 0.45))
        return not _is_myntra_oops_page(p)

    def _wait_for_listing_products(p, timeout_ms: int = 60000) -> bool:
        # Waits for listing cards while auto-recovering from Myntra error pages.
        deadline = time.monotonic() + (timeout_ms / 1000.0)
        while time.monotonic() < deadline:
            if _is_myntra_oops_page(p):
                _recover_from_myntra_oops(p, target_url=p.url, max_attempts=6)
            try:
                n = p.evaluate("() => document.querySelectorAll('li.product-base').length")
                if isinstance(n, int) and n > 0:
                    return True
            except Exception:
                pass
            time.sleep(0.5)
        return False

    def _wait_for_pdp_ready(p, timeout_ms: int = 25000) -> bool:
        deadline = time.monotonic() + (timeout_ms / 1000.0)
        while time.monotonic() < deadline:
            if _is_myntra_oops_page(p):
                _recover_from_myntra_oops(p, target_url=p.url, max_attempts=6)
            try:
                if p.locator("h1.pdp-name, h1.pdp-title, div.pdp-title h1").count() > 0:
                    return True
            except Exception:
                pass
            time.sleep(0.35)
        return False

    def _wait_for_reviews_ready(p, timeout_ms: int = 25000) -> bool:
        deadline = time.monotonic() + (timeout_ms / 1000.0)
        while time.monotonic() < deadline:
            if _is_myntra_oops_page(p):
                _recover_from_myntra_oops(p, target_url=p.url, max_attempts=6)
            try:
                # Any review card/date element or sort control is enough.
                if p.locator("div[itemprop='review'], div.pdp-review, div.review, .reviewCard, .user-review, li.review").count() > 0:
                    return True
                if p.locator(r"text=/Sort\s*by/i").count() > 0:
                    return True
            except Exception:
                pass
            time.sleep(0.35)
        return False

    def _parse_review_date(date_text: str):
        if not date_text:
            return None
        s = re.sub(r"\s+", " ", date_text).strip()
        s = s.replace(",", " ")
        s = re.sub(r"\bSept\b", "Sep", s, flags=re.I)

        now = datetime.now()

        m_tail = re.search(r"(\d{1,2})\s*([A-Za-z]{3,9})\s*(\d{2,4})\s*$", s)
        if m_tail:
            day = int(m_tail.group(1))
            mon = m_tail.group(2)
            if mon.lower() == "sept":
                mon = "Sep"
            year = int(m_tail.group(3))
            if year < 100:
                year += 2000
            for fmt in ("%d %b %Y", "%d %B %Y"):
                try:
                    return datetime.strptime(f"{day} {mon} {year}", fmt)
                except Exception:
                    pass

        if re.fullmatch(r"(?i)today", s):
            return now
        if re.fullmatch(r"(?i)yesterday", s):
            return now - timedelta(days=1)

        m_rel = re.search(r"(?i)\b(\d+)\s*(day|days|month|months|year|years)\s*ago\b", s)
        if m_rel:
            n = int(m_rel.group(1))
            unit = m_rel.group(2).lower()
            if unit.startswith("day"):
                return now - timedelta(days=n)
            if unit.startswith("month"):
                return now - timedelta(days=30 * n)
            if unit.startswith("year"):
                return now - timedelta(days=365 * n)

        m_abs = re.search(r"\b(\d{1,2})\s*([A-Za-z]{3,9})\s*(\d{2,4})\b", s)
        if m_abs:
            day = int(m_abs.group(1))
            mon = m_abs.group(2)
            if mon.lower() == "sept":
                mon = "Sep"
            year = int(m_abs.group(3))
            if year < 100:
                year += 2000
            for fmt in ("%d %b %Y", "%d %B %Y"):
                try:
                    return datetime.strptime(f"{day} {mon} {year}", fmt)
                except Exception:
                    pass

        m_dm = re.search(r"\b(\d{1,2})\s*([A-Za-z]{3,9})\b", s)
        if m_dm:
            day = int(m_dm.group(1))
            mon = m_dm.group(2)
            if mon.lower() == "sept":
                mon = "Sep"
            year = datetime.now().year
            for fmt in ("%d %b %Y", "%d %B %Y"):
                try:
                    dt = datetime.strptime(f"{day} {mon} {year}", fmt)
                    if dt > datetime.now() + timedelta(days=2):
                        dt = datetime(dt.year - 1, dt.month, dt.day)
                    return dt
                except Exception:
                    pass

        m_my = re.search(r"\b([A-Za-z]{3,9})\s*(\d{4})\b", s)
        if m_my:
            mon = m_my.group(1)
            year = int(m_my.group(2))
            for fmt in ("%b %Y", "%B %Y"):
                try:
                    dt = datetime.strptime(f"{mon} {year}", fmt)
                    return datetime(dt.year, dt.month, 1)
                except Exception:
                    pass

        m_num = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b", s)
        if m_num:
            d = int(m_num.group(1))
            m = int(m_num.group(2))
            y = int(m_num.group(3))
            if y < 100:
                y += 2000
            try:
                return datetime(y, m, d)
            except Exception:
                return None

        return None

    def _get_review_counts(dp, product_id=None):
        now = datetime.now()
        window_start = now - timedelta(days=365)
        window_end = now
        years = list(range(2020, 2027))
        year_counts = {y: 0 for y in years}
        # Track standard sizes
        sizes = ["S", "M", "L", "XL", "XXL", "3XL", "4XL"]
        size_counts = {s: 0 for s in sizes}
        rolling_last_12m_count = 0

        if not product_id:
            return "", year_counts, size_counts

        page_num = 1
        size = 40  # Myntra API often caps at 40 or 50
        
        try:
            while True:
                # Use sort=1 for "Most Recent" to ensure we get all reviews chronologically
                api_url = f"https://www.myntra.com/gateway/v1/reviews/product/{product_id}?size={size}&sort=1&rating=0&page={page_num}&includeMetaData=true"
                
                # Fetch via browser context to inherit cookies/headers automatically
                data = dp.evaluate("""async (url) => {
                    try {
                        const response = await fetch(url);
                        if (!response.ok) return { error: `HTTP ${response.status}`, url };
                        return await response.json();
                    } catch (e) {
                        return { error: e.message, url };
                    }
                }""", api_url)
                
                if not data or "error" in data:
                    if data and "error" in data:
                        print(f"  [API Error] {data['error']} for {api_url}")
                    break
                
                reviews = data.get("reviews", [])
                if not reviews or len(reviews) == 0:
                    break
                
                for r in reviews:
                    # Extract size information from styleAttribute
                    style_attrs = r.get("styleAttribute", [])
                    if style_attrs:
                        for attr in style_attrs:
                            if attr.get("name") == "Size bought":
                                sz = str(attr.get("value", "")).upper().strip()
                                if sz in size_counts:
                                    size_counts[sz] += 1
                                break

                    # Myntra API usually provides 'date' as a timestamp in milliseconds
                    raw_date = r.get("date") or r.get("createdOn") or r.get("createdAt") or r.get("updatedAt")
                    dt = None
                    
                    if raw_date:
                        if isinstance(raw_date, (int, float)):
                            dt = datetime.fromtimestamp(raw_date / 1000.0)
                        elif isinstance(raw_date, str):
                            if raw_date.isdigit():
                                dt = datetime.fromtimestamp(int(raw_date) / 1000.0)
                            else:
                                try:
                                    dt = datetime.fromisoformat(raw_date.replace('Z', '+00:00'))
                                except:
                                    dt = _parse_review_date(raw_date)
                    
                    if dt:
                        if window_start <= dt <= window_end:
                            rolling_last_12m_count += 1
                        if dt.year in year_counts:
                            year_counts[dt.year] += 1
                
                # Move to next page
                page_num += 1
                # Safety break to avoid infinite loops
                if page_num > 500: 
                    break
            
            total_yc = sum(year_counts.values())
            print(f"  [API] Finished fetching. Total reviews processed: {total_yc}")
                
        except Exception as e:
            print(f"Error fetching reviews via API for product {product_id}: {e}")
            
        return rolling_last_12m_count, year_counts, size_counts


    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 768},
        )

        def should_block(req):
            rt = req.resource_type
            u = req.url
            if rt in ["image", "font", "media", "stylesheet"]:
                return True
            if re.search(r"(doubleclick|google-analytics|adservice|facebook|hotjar|segment|optimizely|ads|tracking)", u, re.I):
                return True
            return False

        context.route("**/*", lambda route: route.abort() if should_block(route.request) else route.continue_())

        def _new_listing_page():
            p2 = context.new_page()
            p2.set_default_timeout(20000)
            return p2

        def _new_detail_page():
            p2 = context.new_page()
            p2.set_default_timeout(15000)
            return p2

        page = _new_listing_page()
        detail_page = _new_detail_page()

        def load_page_with_retries(url, retries=3, delay=2):
            nonlocal page
            for attempt in range(retries):
                try:
                    if page.is_closed():
                        page = _new_listing_page()

                    resp = page.goto(url, timeout=90000, wait_until="domcontentloaded")
                    try:
                        page.wait_for_load_state("networkidle", timeout=45000)
                    except Exception:
                        pass

                    try:
                        if resp is not None and getattr(resp, "status", None) is not None and resp.status >= 400:
                            raise RuntimeError(f"HTTP {resp.status}")
                    except Exception:
                        pass

                    # Auto-recover if Myntra shows the "Oops...Refresh" page.
                    if _is_myntra_oops_page(page):
                        _recover_from_myntra_oops(page, target_url=url, max_attempts=8)
                        # Re-wait for the listing to be ready after recovery.
                        try:
                            page.wait_for_load_state("networkidle", timeout=45000)
                        except Exception:
                            pass

                    if not _wait_for_listing_products(page, timeout_ms=60000):
                        raise RuntimeError("Timed out waiting for product cards")

                    # If we still landed on an error page (or got redirected), force a retry.
                    if _is_myntra_oops_page(page):
                        raise RuntimeError("Myntra returned 'Oops! Something went wrong' page")

                    time.sleep(0.5 + random.uniform(0.0, 0.3))
                    return True
                except Exception as e:
                    print(f"Load attempt {attempt + 1} failed: {e}")
                    try:
                        if not page.is_closed():
                            page.close()
                    except Exception:
                        pass
                    try:
                        page = _new_listing_page()
                    except Exception:
                        pass
                    time.sleep(delay)
            return False

        try:
            max_pages_env = os.environ.get("MYNTRA_MAX_PAGES", "")
            max_pages = int(max_pages_env) if max_pages_env.strip() else 0
            if max_pages < 0:
                max_pages = 0
        except Exception:
            max_pages = 0

        for list_url in (urls or []):
            list_url = (list_url or "").strip()
            if not list_url:
                continue

            print(f"\n Starting URL: {list_url}")
            if not load_page_with_retries(list_url):
                print("Failed to load URL after retries. Skipping.")
                continue

            page_number = 1
            while True:
                print(f" Scraping page {page_number}...")
                products = page.query_selector_all("li.product-base")

                try:
                    max_products = int(os.environ.get("MYNTRA_MAX_PRODUCTS", "0") or "0")
                except Exception:
                    max_products = 0
                if max_products and max_products > 0:
                    products = (products or [])[:max_products]

                for product in products:
                    try:
                        name = product.query_selector("h4.product-product").inner_text().strip()
                    except Exception:
                        name = ""

                    try:
                        brand = product.query_selector("h3.product-brand").inner_text().strip()
                    except Exception:
                        brand = ""

                    try:
                        rating = product.query_selector("div.product-ratingsContainer span").inner_text().strip()
                    except Exception:
                        rating = "0"

                    try:
                        rating_count = product.query_selector("div.product-ratingsContainer div.product-ratingsCount").inner_text().strip().replace("|", "").strip()
                    except Exception:
                        rating_count = "0"

                    try:
                        price_el = product.query_selector("span.product-discountedPrice") or product.query_selector("span.product-price") or product.query_selector("div.product-price span")
                        price_txt = price_el.inner_text().strip() if price_el else ""
                        if price_txt:
                            mpr = re.search(r'(₹|Rs\.?)[\s]*([\d,]+)', price_txt)
                            price = (mpr.group(1) + " " + mpr.group(2)) if mpr else price_txt
                        else:
                            price = ""
                    except Exception:
                        price = ""

                    link = ""
                    try:
                        buy_anchor = product.query_selector('a[href*="/buy"]')
                        if buy_anchor:
                            link = buy_anchor.get_attribute('href') or ""
                        else:
                            anchors = product.query_selector_all('a')
                            for a in anchors or []:
                                href = (a.get_attribute('href') or a.get_attribute('data-href') or a.get_attribute('data-url') or "").strip()
                                if not href:
                                    continue
                                if re.search(r'/\d{6,}', href):
                                    link = href
                                    break
                            if not link and anchors:
                                link = anchors[0].get_attribute('href') or ""

                        if link:
                            if not link.startswith("http"):
                                if not link.startswith("/"):
                                    link = "/" + link
                                link = "https://www.myntra.com" + link
                            if re.search(r'/\d{6,}(?:$|\?|/)', link) and not re.search(r'/buy(?:$|\?|/)', link):
                                link = re.sub(r'(\/\d{6,})(?:\/)?(?=$|\?|/)', r'\1/buy', link)
                    except Exception:
                        link = ""

                    long_name = ""
                    product_rating = rating
                    reviews_count = rating_count
                    last_12m_reviews_count = ""
                    ycounts = {2020: "", 2021: "", 2022: "", 2023: "", 2024: "", 2025: "", 2026: ""}
                    scounts = {"S": "", "M": "", "L": "", "XL": "", "XXL": "", "3XL": "", "4XL": ""}

                    if link:
                        nonlocal_detail_page = False
                        try:
                            # Ensure the detail page is usable; recover if it was closed.
                            if detail_page.is_closed():
                                nonlocal_detail_page = True
                        except Exception:
                            nonlocal_detail_page = True
                        if nonlocal_detail_page:
                            try:
                                detail_page = _new_detail_page()
                            except Exception:
                                detail_page = _new_detail_page()

                        dp = detail_page
                        try:
                            nav_ok = False
                            for _nav_attempt in range(2):
                                try:
                                    dp.goto(link, timeout=45000, wait_until="domcontentloaded")
                                    dp.wait_for_selector("body", timeout=10000)
                                    nav_ok = True
                                    break
                                except Exception as e:
                                    msg = str(e) if e is not None else ""
                                    if "Target page" in msg or "closed" in msg.lower():
                                        try:
                                            detail_page = _new_detail_page()
                                            dp = detail_page
                                            continue
                                        except Exception:
                                            break
                                    raise

                            if not nav_ok:
                                raise RuntimeError("Failed to open product detail page")

                            # Auto-recover if Myntra shows the "Oops...Refresh" page.
                            if _is_myntra_oops_page(dp):
                                _recover_from_myntra_oops(dp, target_url=link, max_attempts=8)
                                try:
                                    dp.wait_for_selector("body", timeout=10000)
                                except Exception:
                                    pass
                            _wait_for_pdp_ready(dp, timeout_ms=25000)

                            try:
                                rt_el = (
                                    dp.query_selector("div.index-overallRating")
                                    or dp.query_selector("span.index-overallRating")
                                    or dp.query_selector("div.pdp-ratings span")
                                    or dp.query_selector("div.pdp-product-rating span")
                                )
                                rt_txt = rt_el.inner_text().strip() if rt_el else ""
                                if rt_txt:
                                    mrt = re.search(r"(\d+(?:\.\d+)?)", rt_txt)
                                    product_rating = mrt.group(1) if mrt else rt_txt
                            except Exception:
                                pass

                            try:
                                if not price:
                                    lp = dp.query_selector("span.pdp-price") or dp.query_selector("span.pdp-offers-price") or dp.query_selector("div.pdp-price span") or dp.query_selector("span[class*='pdp-price']")
                                    cp = lp.inner_text().strip() if lp else ""
                                    if cp:
                                        mpr = re.search(r'(₹|Rs\.?)[\s]*([\d,]+)', cp)
                                        price = (mpr.group(1) + " " + mpr.group(2)) if mpr else cp
                            except Exception:
                                pass

                            try:
                                rc_el = dp.query_selector("div.pdp-reviews-count") or dp.query_selector("div.product-ratingsCount") or dp.query_selector("span.pdp-reviews-count")
                                rc_txt = rc_el.inner_text().strip() if rc_el else ""
                                if rc_txt:
                                    mrc = re.search(r"(\d+[\,\d]*)", rc_txt)
                                    reviews_count = mrc.group(1) if mrc else rc_txt
                            except Exception:
                                pass

                            try:
                                ln_el = (
                                    dp.query_selector("h1.pdp-name")
                                    or dp.query_selector("h1.pdp-title")
                                    or dp.query_selector("div.pdp-title h1")
                                    or dp.query_selector("div.pdp-title")
                                )
                                long_name = (ln_el.inner_text().strip() if ln_el else "")
                                if not long_name:
                                    og = dp.query_selector("meta[property='og:title']")
                                    long_name = (og.get_attribute("content").strip() if og and og.get_attribute("content") else "")
                                if not long_name and brand and name:
                                    long_name = f"{brand} {name}".strip()
                            except Exception:
                                pass

                            try:
                                m_pid = None
                                m = re.search(r"/(\d{6,})(?:/buy)?(?:$|\?|/)", dp.url or "")
                                m_pid = m.group(1) if m else None
                                c12, yc, sc = _get_review_counts(dp, product_id=m_pid)
                                last_12m_reviews_count = str(c12) if c12 != "" else ""
                                if isinstance(yc, dict):
                                    ycounts = yc
                                if isinstance(sc, dict):
                                    scounts = sc
                            except Exception:
                                pass

                            try:
                                link = dp.url or link
                            except Exception:
                                pass

                            time.sleep(0.2 + random.uniform(0.05, 0.15))
                        except Exception:
                            pass

                    if not long_name:
                        long_name = f"{brand} {name}".strip() if (brand and name) else (name or "")

                    row = [
                        (brand or "").replace('\n', ' ').strip(),
                        (long_name or "").replace('\n', ' ').strip(),
                        (price or "").strip(),
                        (product_rating or "").strip(),
                        (reviews_count or "").strip(),
                        (link or "").strip(),
                        (str(last_12m_reviews_count) if last_12m_reviews_count is not None else ""),
                        (str(ycounts.get(2020, "")) if ycounts.get(2020, "") != "" else ""),
                        (str(ycounts.get(2021, "")) if ycounts.get(2021, "") != "" else ""),
                        (str(ycounts.get(2022, "")) if ycounts.get(2022, "") != "" else ""),
                        (str(ycounts.get(2023, "")) if ycounts.get(2023, "") != "" else ""),
                        (str(ycounts.get(2024, "")) if ycounts.get(2024, "") != "" else ""),
                        (str(ycounts.get(2025, "")) if ycounts.get(2025, "") != "" else ""),
                        (str(ycounts.get(2026, "")) if ycounts.get(2026, "") != "" else ""),
                        (str(scounts.get("S", "")) if scounts.get("S", "") != "" else ""),
                        (str(scounts.get("M", "")) if scounts.get("M", "") != "" else ""),
                        (str(scounts.get("L", "")) if scounts.get("L", "") != "" else ""),
                        (str(scounts.get("XL", "")) if scounts.get("XL", "") != "" else ""),
                        (str(scounts.get("XXL", "")) if scounts.get("XXL", "") != "" else ""),
                        (str(scounts.get("3XL", "")) if scounts.get("3XL", "") != "" else ""),
                        (str(scounts.get("4XL", "")) if scounts.get("4XL", "") != "" else ""),
                        scrape_timestamp,
                    ]

                    results.append(dict(zip(csv_headers, row)))

                    try:
                        with open(csv_name, 'a', newline='', encoding='utf-8') as f:
                            csv.writer(f).writerow(row)
                    except Exception as e:
                        print(f"Warning: could not write to CSV ({csv_name}): {e}")

                if max_pages and page_number >= max_pages:
                    print(f" Stopping after page {page_number} as requested (MYNTRA_MAX_PAGES={max_pages}).")
                    break

                try:
                    next_button = page.query_selector("li.pagination-next")
                except Exception as e:
                    print(f"Pagination unavailable or page closed: {e}")
                    break

                if next_button is None or "disabled" in (next_button.get_attribute("class") or ""):
                    print(" No more pages to scrape.")
                    break

                max_click_retries = 3
                click_attempt = 0
                while click_attempt < max_click_retries:
                    try:
                        first_product = page.query_selector("li.product-base")
                        first_product_html = first_product.inner_html() if first_product else ""
                        next_button.scroll_into_view_if_needed()
                        next_button.click()
                        page.wait_for_load_state("domcontentloaded")

                        page.wait_for_function(
                            "firstHTML => document.querySelector('li.product-base') && document.querySelector('li.product-base').innerHTML !== firstHTML",
                            arg=first_product_html,
                            timeout=20000,
                        )
                        page_number += 1
                        time.sleep(0.4 + random.uniform(0.1, 0.3))
                        break
                    except Exception as e:
                        click_attempt += 1
                        print(f"Attempt {click_attempt} to navigate to next page failed: {e}")
                        time.sleep(0.4)
                else:
                    print(" Failed to navigate to next page after retries. Stopping.")
                    break

        try:
            browser.close()
        except Exception:
            pass

    try:
        if os.path.exists(xlsx_name) and os.path.getsize(xlsx_name) > 0:
            wb = load_workbook(xlsx_name)
            ws = wb.active
            existing_headers = [c.value for c in (ws[1] if ws.max_row >= 1 else [])]
            if existing_headers != csv_headers:
                # If the file exists but headers differ, do not overwrite; create headers in a new empty workbook.
                wb = Workbook()
                ws = wb.active
                ws.append(csv_headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(csv_headers)

        for item in results:
            ws.append([item.get(h, "") for h in csv_headers])

        wb.save(xlsx_name)
    except Exception as e:
        print(f"Warning: could not write to Excel ({xlsx_name}): {e}")

    print(f"\n Scraping completed. {len(results)} products appended to {xlsx_name}.")


if __name__ == "__main__":
    scrape_mns_myntra()
